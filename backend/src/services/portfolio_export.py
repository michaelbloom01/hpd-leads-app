"""Read-only portfolio building/contact export helpers."""
from __future__ import annotations

import csv
import re
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from src.services.contact_roster import get_building_contacts


VENTURE_KEYS = {
    "VENTURENYPROPERTYMANAGEMENT",
    "VENTURENYPROPERTYMANAGEMENTLLC",
}


def normalize_company_key(value: str | None) -> str:
    return re.sub(r"[^A-Z0-9]", "", (value or "").upper())


def json_default(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def flatten_contact_rows(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for building in records:
        contacts = building.get("contacts") or []
        if not contacts:
            rows.append({
                "bbl": building.get("bbl"),
                "address": building.get("address"),
                "borough": building.get("borough"),
                "zip_code": building.get("zip_code"),
                "unit_count": building.get("unit_count"),
                "churn_score": building.get("churn_score"),
                "churn_category": building.get("churn_category"),
                "management_company": building.get("management_company"),
                "corporate_owner": building.get("corporate_owner"),
                "dos_contacts_status": building.get("dos_contacts_status"),
                "contact_name": None,
                "contact_role": None,
                "contact_source": None,
                "contact_updated": None,
                "contact_address": None,
                "contact_confidence": None,
                "contact_source_record_id": None,
                "contact_source_url": None,
                "board_role": None,
                "is_decision_maker": None,
            })
            continue
        for contact in contacts:
            rows.append({
                "bbl": building.get("bbl"),
                "address": building.get("address"),
                "borough": building.get("borough"),
                "zip_code": building.get("zip_code"),
                "unit_count": building.get("unit_count"),
                "churn_score": building.get("churn_score"),
                "churn_category": building.get("churn_category"),
                "management_company": building.get("management_company"),
                "corporate_owner": building.get("corporate_owner"),
                "dos_contacts_status": building.get("dos_contacts_status"),
                "contact_name": contact.get("name"),
                "contact_role": contact.get("role"),
                "contact_source": contact.get("source"),
                "contact_updated": contact.get("as_of_date"),
                "contact_address": contact.get("address"),
                "contact_confidence": contact.get("confidence_hint") or "--",
                "contact_source_record_id": contact.get("source_record_id"),
                "contact_source_url": contact.get("source_url"),
                "board_role": contact.get("board_role"),
                "is_decision_maker": contact.get("is_decision_maker"),
            })
    return rows


async def build_portfolio_export(
    *,
    session: AsyncSession,
    company: str,
    lead_id: str | None = None,
) -> dict[str, Any]:
    company_key = normalize_company_key(company)
    keys = sorted(VENTURE_KEYS if company_key in VENTURE_KEYS else {company_key})

    params: dict[str, Any] = {"keys": keys}
    lead_filter = ""
    if lead_id:
        params["lead_id"] = lead_id
        lead_filter = """
            OR EXISTS (
                SELECT 1
                FROM building_management bm
                WHERE bm.bbl = b.bbl
                  AND bm.lead_id = :lead_id
                  AND bm.is_current = true
            )
        """

    result = await session.execute(
        text(f"""
            WITH agent_portfolio AS (
                SELECT DISTINCT bc.bbl
                FROM building_contacts bc
                WHERE bc.contact_type = 'Agent'
                  AND regexp_replace(upper(coalesce(bc.corporation_name, '')), '[^A-Z0-9]', '', 'g')
                      = ANY(:keys)
            )
            SELECT
                b.bbl, b.bin, b.address, b.borough, b.block, b.lot, b.zip_code,
                b.building_class, b.building_type, b.unit_count, b.year_built,
                b.assessed_value, b.churn_score, b.churn_category,
                b.churn_breakdown, b.key_signal, b.outreach_status,
                b.last_scored_at, b.updated_at
            FROM buildings b
            WHERE b.bbl IN (SELECT bbl FROM agent_portfolio)
            {lead_filter}
            ORDER BY b.address ASC, b.bbl ASC
        """),
        params,
    )
    building_rows = [dict(row._mapping) for row in result]

    records: list[dict[str, Any]] = []
    source_counts: dict[str, int] = {}
    total_contacts = 0
    dos_status_counts: dict[str, int] = {}
    for building in building_rows:
        contacts, meta = await get_building_contacts(
            session=session,
            bbl=str(building["bbl"]),
            building_address=building.get("address"),
        )
        for contact in contacts:
            source = str(contact.get("source") or "Unknown")
            source_counts[source] = source_counts.get(source, 0) + 1
        total_contacts += len(contacts)
        dos_status = str(meta.get("dos_contacts_status") or "unknown")
        dos_status_counts[dos_status] = dos_status_counts.get(dos_status, 0) + 1
        records.append({
            **building,
            "management_company": meta.get("management_company"),
            "corporate_owner": meta.get("corporate_owner"),
            "dos_contacts_status": dos_status,
            "dos_contacts_is_stale": bool(meta.get("dos_contacts_is_stale")),
            "dos_contacts_last_refreshed_at": meta.get("dos_contacts_last_refreshed_at"),
            "contacts": contacts,
            "contact_count": len(contacts),
            "contact_source_count": len({c.get("source") for c in contacts if c.get("source")}),
        })

    return {
        "run_type": "portfolio_building_contacts_export",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "dry_run": True,
        "mutations_planned": 0,
        "company": company,
        "company_keys": keys,
        "portfolio_definition": (
            "Buildings whose HPD Registration Agent corporation name normalizes to the company key"
            + (" plus explicit lead_id current links." if lead_id else ".")
        ),
        "lead_id": lead_id,
        "building_count": len(records),
        "unit_count": sum(int(b.get("unit_count") or 0) for b in records),
        "contact_count": total_contacts,
        "contact_source_counts": dict(sorted(source_counts.items())),
        "dos_status_counts": dict(sorted(dos_status_counts.items())),
        "records": records,
        "contact_rows": flatten_contact_rows(records),
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _excel_value(value: Any) -> str | int | float | bool | None:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _set_sheet_widths(sheet, rows: list[list[Any]], *, max_width: int = 54) -> None:
    max_columns = max((len(row) for row in rows), default=1)
    for column_index in range(1, max_columns + 1):
        letter = get_column_letter(column_index)
        values = [row[column_index - 1] for row in rows if len(row) >= column_index]
        width = min(max(len(str(value or "")) + 2 for value in values[:200]), max_width)
        sheet.column_dimensions[letter].width = max(width, 10)


def _add_table_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    data_rows: list[list[Any]],
) -> None:
    sheet = workbook.create_sheet(title)
    rows = [headers, *data_rows]
    for row in rows:
        sheet.append([_excel_value(value) for value in row])

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _set_sheet_widths(sheet, rows)


def _contact_source_counts(contact_rows: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in contact_rows:
        source = str(row.get("contact_source") or "No contact")
        counts[source] = counts.get(source, 0) + 1
    return dict(sorted(counts.items()))


def _building_source_summary(record: dict[str, Any]) -> str:
    sources = sorted({
        str(contact.get("source"))
        for contact in record.get("contacts") or []
        if contact.get("source")
    })
    return ", ".join(sources) if sources else "No contact source loaded"


def _gap_rows(records: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for record in records:
        contacts = record.get("contacts") or []
        sources = {contact.get("source") for contact in contacts if contact.get("source")}
        if not contacts:
            rows.append([
                record.get("bbl"),
                record.get("address"),
                "no_contacts",
                "No building-level contacts are loaded for this portfolio building.",
                "Do not use for direct outreach until a contact source is loaded.",
            ])
        if record.get("dos_contacts_status") in {"not_loaded", "missing", "unknown"}:
            rows.append([
                record.get("bbl"),
                record.get("address"),
                "dos_not_loaded",
                "DOS corporate officer contacts are not loaded for this building.",
                "Load/confirm DOS before treating owner/officer relationships as complete.",
            ])
        if record.get("dos_contacts_is_stale"):
            rows.append([
                record.get("bbl"),
                record.get("address"),
                "dos_stale",
                "DOS contact cache is marked stale for this building.",
                "Refresh DOS before using stale officer contact data.",
            ])
        if contacts and len(sources) <= 1:
            rows.append([
                record.get("bbl"),
                record.get("address"),
                "single_source_contacts",
                "Contacts are present, but only one independent source family is represented.",
                "Use as sourced lead context, not as verified multi-source truth.",
            ])
    return rows


def build_portfolio_workbook(payload: dict[str, Any]) -> bytes:
    """Render a read-only portfolio export payload as a native Excel workbook."""
    records = payload.get("records") or []
    contact_rows = payload.get("contact_rows") or []
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    summary_rows = [
        ["Double Edge Portfolio Contacts Export", ""],
        ["Company", payload.get("company")],
        ["Generated at", payload.get("generated_at")],
        ["Read-only posture", "dry_run=true; mutations_planned=0"],
        ["Portfolio definition", payload.get("portfolio_definition")],
        ["Buildings", payload.get("building_count")],
        ["Units", payload.get("unit_count")],
        ["Contact rows", payload.get("contact_count")],
        [
            "Truth caveat",
            (
                "This workbook reports sourced contact evidence for outreach/research. "
                "It does not mark single-source claims verified."
            ),
        ],
        ["Company keys", ", ".join(payload.get("company_keys") or [])],
    ]
    for row in summary_rows:
        summary.append([_excel_value(value) for value in row])
    summary["A1"].font = Font(bold=True, size=16, color="0F172A")
    summary["A1"].fill = PatternFill("solid", fgColor="DFF6DD")
    summary.merge_cells("A1:B1")
    for cell in summary["A"]:
        cell.font = Font(bold=True, color="334155")
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _set_sheet_widths(summary, summary_rows, max_width=72)

    source_rows = [
        [source, count, "sampled_export_contact_rows"]
        for source, count in _contact_source_counts(contact_rows).items()
    ]
    source_rows.extend([
        [f"DOS status: {status}", count, "building_dos_contact_cache_status"]
        for status, count in (payload.get("dos_status_counts") or {}).items()
    ])
    _add_table_sheet(
        workbook,
        "Source Coverage",
        ["source_or_status", "row_or_building_count", "coverage_basis"],
        source_rows,
    )

    building_rows = [[
        record.get("bbl"),
        record.get("bin"),
        record.get("address"),
        record.get("borough"),
        record.get("zip_code"),
        record.get("unit_count"),
        record.get("year_built"),
        record.get("building_class"),
        record.get("building_type"),
        record.get("churn_score"),
        record.get("churn_category"),
        record.get("management_company"),
        record.get("corporate_owner"),
        record.get("dos_contacts_status"),
        record.get("contact_count"),
        record.get("contact_source_count"),
        _building_source_summary(record),
        record.get("last_scored_at"),
        record.get("updated_at"),
    ] for record in records]
    _add_table_sheet(
        workbook,
        "Buildings",
        [
            "bbl",
            "bin",
            "address",
            "borough",
            "zip_code",
            "unit_count",
            "year_built",
            "building_class",
            "building_type",
            "churn_score",
            "churn_category",
            "management_company",
            "corporate_owner",
            "dos_contacts_status",
            "contact_count",
            "contact_source_count",
            "contact_sources",
            "last_scored_at",
            "updated_at",
        ],
        building_rows,
    )

    contact_headers = list(contact_rows[0].keys()) if contact_rows else [
        "bbl",
        "address",
        "borough",
        "zip_code",
        "unit_count",
        "churn_score",
        "churn_category",
        "management_company",
        "corporate_owner",
        "dos_contacts_status",
        "contact_name",
        "contact_role",
        "contact_source",
        "contact_updated",
        "contact_address",
        "contact_confidence",
        "contact_source_record_id",
        "contact_source_url",
        "board_role",
        "is_decision_maker",
    ]
    _add_table_sheet(
        workbook,
        "Contacts",
        contact_headers,
        [[row.get(header) for header in contact_headers] for row in contact_rows],
    )
    _add_table_sheet(
        workbook,
        "Gaps",
        ["bbl", "address", "gap_type", "why_it_matters", "safe_action"],
        _gap_rows(records),
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def write_xlsx(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(build_portfolio_workbook(payload))
